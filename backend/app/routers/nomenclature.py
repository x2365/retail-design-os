"""Номенклатура / штрих-коды задачи (П3): CRUD, статусы, выгрузка в Excel для РЦ."""

from __future__ import annotations

import io

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from .. import models, schemas, security, serializers
from ..database import get_db

router = APIRouter(tags=["nomenclature"])

ReadDep = Depends(security.get_current_user)
WriteDep = Depends(security.require_roles(models.Role.manager))


def _get_task(db: Session, code: str) -> models.Task:
    task = db.scalar(select(models.Task).where(models.Task.code == code))
    if not task:
        raise HTTPException(404, f"Task {code} not found")
    return task


def _items(db: Session, task_id: int):
    return db.scalars(
        select(models.NomenclatureItem)
        .options(selectinload(models.NomenclatureItem.task))
        .where(models.NomenclatureItem.task_id == task_id)
        .order_by(models.NomenclatureItem.id)
    ).all()


@router.get("/tasks/{code}/nomenclature", response_model=list[schemas.NomenclatureItemOut])
def list_nomenclature(code: str, db: Session = Depends(get_db), _user: models.User = ReadDep):
    task = _get_task(db, code)
    return [serializers.nomenclature_to_out(n) for n in _items(db, task.id)]


@router.post(
    "/tasks/{code}/nomenclature", response_model=schemas.NomenclatureItemOut, status_code=201
)
def create_nomenclature(
    code: str,
    payload: schemas.NomenclatureItemCreate,
    db: Session = Depends(get_db),
    _user: models.User = WriteDep,
):
    task = _get_task(db, code)
    n = models.NomenclatureItem(
        task_id=task.id,
        sku=payload.sku,
        barcode=payload.barcode,
        name=payload.name or task.name,
        qty=payload.qty,
    )
    db.add(n)
    db.commit()
    reloaded = db.scalar(
        select(models.NomenclatureItem)
        .options(selectinload(models.NomenclatureItem.task))
        .where(models.NomenclatureItem.id == n.id)
    )
    assert reloaded is not None
    return serializers.nomenclature_to_out(reloaded)


@router.patch("/nomenclature/{item_id}", response_model=schemas.NomenclatureItemOut)
def update_nomenclature(
    item_id: int,
    payload: schemas.NomenclatureItemUpdate,
    db: Session = Depends(get_db),
    _user: models.User = WriteDep,
):
    n = db.scalar(
        select(models.NomenclatureItem)
        .options(selectinload(models.NomenclatureItem.task))
        .where(models.NomenclatureItem.id == item_id)
    )
    if not n:
        raise HTTPException(404, "Позиция не найдена")
    data = payload.model_dump(exclude_unset=True)
    if "status" in data:
        try:
            n.status = models.NomenclatureStatus(data.pop("status"))
        except ValueError as exc:
            raise HTTPException(422, "Недопустимый статус") from exc
    for k, v in data.items():
        setattr(n, k, v)
    db.commit()
    db.refresh(n)
    return serializers.nomenclature_to_out(n)


@router.delete("/nomenclature/{item_id}", status_code=204)
def delete_nomenclature(item_id: int, db: Session = Depends(get_db), _user: models.User = WriteDep):
    n = db.get(models.NomenclatureItem, item_id)
    if not n:
        raise HTTPException(404, "Позиция не найдена")
    db.delete(n)
    db.commit()


@router.post(
    "/tasks/{code}/nomenclature/send-to-rc", response_model=list[schemas.NomenclatureItemOut]
)
def send_to_rc(code: str, db: Session = Depends(get_db), _user: models.User = WriteDep):
    """Перевести все позиции задачи в статус «Направлено на РЦ»."""
    task = _get_task(db, code)
    items = _items(db, task.id)
    for n in items:
        if n.status == models.NomenclatureStatus.draft:
            n.status = models.NomenclatureStatus.sent_to_rc
    db.commit()
    return [serializers.nomenclature_to_out(n) for n in _items(db, task.id)]


@router.get("/tasks/{code}/nomenclature.xlsx")
def export_nomenclature_xlsx(
    code: str, db: Session = Depends(get_db), _user: models.User = ReadDep
):
    """Excel-шаблон ШК для отправки на РЦ."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    task = _get_task(db, code)
    items = _items(db, task.id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Номенклатура"
    headers = ["SKU", "Штрих-код", "Наименование", "Кол-во", "Статус"]
    ws.append(headers)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="5B6AF0")
    for cell in ws[1]:
        cell.font = head_font
        cell.fill = head_fill
    for n in items:
        ws.append(
            [
                n.sku,
                n.barcode,
                n.name,
                n.qty,
                models.NOMENCLATURE_STATUS_LABELS.get(n.status, n.status.value),
            ]
        )
    widths = [18, 18, 40, 10, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"nomenclature_{task.code}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
